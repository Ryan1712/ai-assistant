import uuid
from datetime import date, timedelta

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select

from app.config import Settings
from app.models import (
    Project, Report, Role, Task, TaskAssignee, TaskStatus, TaskUpdate, User, Workspace,
)
from app.services import report_service


def test_storage_dir_setting_default():
    assert Settings().storage_dir == "./storage/reports"


@pytest.mark.asyncio
async def test_report_model_roundtrip(db_session):
    ws = Workspace(name="A")
    db_session.add(ws)
    await db_session.flush()
    ceo = User(workspace_id=ws.id, email="c@a.vn", password_hash="x", full_name="C",
              role=Role.ceo, is_root=True)
    db_session.add(ceo)
    await db_session.flush()
    report = Report(workspace_id=ws.id, requested_by=ceo.id,
                    filters={"status": "done"}, summary={"total": 0},
                    file_path=f"{ws.id}/x.xlsx")
    db_session.add(report)
    await db_session.commit()

    fetched = await db_session.get(Report, report.id)
    assert fetched.kind == "task_summary"
    assert fetched.filters == {"status": "done"}
    assert fetched.summary == {"total": 0}
    assert fetched.created_at is not None


async def _seed(db):
    ws = Workspace(name="A")
    db.add(ws)
    await db.flush()
    ceo = User(workspace_id=ws.id, email="ceo@a.vn", password_hash="x", full_name="Sep",
              role=Role.ceo, is_root=True)
    emp = User(workspace_id=ws.id, email="e@a.vn", password_hash="x", full_name="Nhan Vien",
              role=Role.employee)
    db.add_all([ceo, emp])
    await db.flush()
    project = Project(workspace_id=ws.id, name="Website", created_by=ceo.id)
    db.add(project)
    await db.flush()
    return ws, ceo, emp, project


async def _task(db, ws, project, ceo, title, status=TaskStatus.todo, percent=0):
    task = Task(workspace_id=ws.id, project_id=project.id, title=title,
                status=status, percent=percent, created_by=ceo.id)
    db.add(task)
    await db.flush()
    return task


@pytest.mark.asyncio
async def test_generate_report_writes_xlsx_and_summary(db_session, storage_dir):
    ws, ceo, emp, project = await _seed(db_session)
    t1 = await _task(db_session, ws, project, ceo, "Lam trang chu",
                     status=TaskStatus.in_progress, percent=40)
    await _task(db_session, ws, project, ceo, "Viet tai lieu", status=TaskStatus.done,
                percent=100)
    db_session.add(TaskAssignee(workspace_id=ws.id, task_id=t1.id, user_id=emp.id))
    db_session.add(TaskUpdate(workspace_id=ws.id, task_id=t1.id, author_id=emp.id,
                              content="da xong 40%", percent=40))
    await db_session.commit()

    result = await report_service.generate_report(db_session, ceo)

    assert result["row_count"] == 2
    assert result["summary"] == {"total": 2, "todo": 0, "in_progress": 1,
                                 "blocked": 0, "done": 1}
    report = await db_session.get(Report, uuid.UUID(result["report_id"]))
    assert report.workspace_id == ws.id
    assert report.requested_by == ceo.id
    assert report.summary == result["summary"]

    path = storage_dir / str(ws.id) / f"{result['report_id']}.xlsx"
    assert path.is_file()
    assert report.file_path == f"{ws.id}/{result['report_id']}.xlsx"
    sheet = load_workbook(path).active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("Tên task", "Project", "Trạng thái", "% hoàn thành",
                       "Người phụ trách", "Cập nhật mới nhất", "Deadline")
    assert len(rows) == 3
    row1 = next(r for r in rows[1:] if r[0] == "Lam trang chu")
    assert row1[1] == "Website"
    assert row1[2] == "in_progress"
    assert row1[3] == 40
    assert row1[4] == "Nhan Vien"
    assert "da xong 40%" in row1[5]


@pytest.mark.asyncio
async def test_generate_report_forbidden_for_non_ceo(db_session, storage_dir):
    ws, ceo, emp, project = await _seed(db_session)
    await db_session.commit()
    with pytest.raises(HTTPException) as exc:
        await report_service.generate_report(db_session, emp)
    assert exc.value.status_code == 403


@pytest.mark.asyncio
async def test_generate_report_unknown_project_404(db_session, storage_dir):
    ws, ceo, emp, project = await _seed(db_session)
    await db_session.commit()
    with pytest.raises(HTTPException) as exc:
        await report_service.generate_report(db_session, ceo, project_id=uuid.uuid4())
    assert exc.value.status_code == 404


@pytest.mark.asyncio
async def test_generate_report_unknown_assignee_404(db_session, storage_dir):
    ws, ceo, emp, project = await _seed(db_session)
    await db_session.commit()
    with pytest.raises(HTTPException) as exc:
        await report_service.generate_report(db_session, ceo, assignee_id=uuid.uuid4())
    assert exc.value.status_code == 404


@pytest.mark.asyncio
async def test_generate_report_filters(db_session, storage_dir):
    ws, ceo, emp, project = await _seed(db_session)
    other = Project(workspace_id=ws.id, name="App", created_by=ceo.id)
    db_session.add(other)
    await db_session.flush()
    t1 = await _task(db_session, ws, project, ceo, "T1", status=TaskStatus.done)
    await _task(db_session, ws, other, ceo, "T2", status=TaskStatus.todo)
    db_session.add(TaskAssignee(workspace_id=ws.id, task_id=t1.id, user_id=emp.id))
    await db_session.commit()

    by_project = await report_service.generate_report(db_session, ceo,
                                                      project_id=project.id)
    assert by_project["row_count"] == 1

    by_status = await report_service.generate_report(db_session, ceo,
                                                     status=TaskStatus.done)
    assert by_status["row_count"] == 1

    by_assignee = await report_service.generate_report(db_session, ceo,
                                                       assignee_id=emp.id)
    assert by_assignee["row_count"] == 1
    assert by_assignee["filters_applied"]["assignee_id"] == str(emp.id)

    future = date.today() + timedelta(days=7)
    by_date = await report_service.generate_report(db_session, ceo, date_from=future)
    assert by_date["row_count"] == 0


@pytest.mark.asyncio
async def test_generate_report_empty_still_creates_file(db_session, storage_dir):
    ws, ceo, emp, project = await _seed(db_session)
    await db_session.commit()
    result = await report_service.generate_report(db_session, ceo,
                                                  status=TaskStatus.blocked)
    assert result["summary"]["total"] == 0
    path = storage_dir / str(ws.id) / f"{result['report_id']}.xlsx"
    assert path.is_file()
    sheet = load_workbook(path).active
    assert len(list(sheet.iter_rows())) == 1  # chỉ header
